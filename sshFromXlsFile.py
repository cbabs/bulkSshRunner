import netmiko
import time
import getpass
from openpyxl import load_workbook
import re


# TS for file where all output is logged
timestamp = time.ctime().replace(':', '.')


hostCmdsList = []

def createDictFromXls():

    # Load workbook and sheet
    wb = load_workbook(filename = 'hostCmnd.xlsx')  
    sheet = wb.active
    
    for row in sheet.iter_rows():
        
        #hostCol = str(row[0].value)
        hostCol = str(row[0].value)
        cmdCol = str(row[1].value)
        cmdCol = cmdCol.replace("\\n", "\n",)
        
        # Skip row if begins with ###
        if re.match(r"^###", hostCol): continue
        
        hostCmdsList.append({hostCol: cmdCol})
        
        

def getCreds():

    while True:
        sshUser = input('\n\nEnter username: ')
        sshPass = getpass.getpass('Enter password: ')
        devType = input('''Device Type? (Enter "?" for list.  Hit enter for generic termserver)
SSH Device Type: ''')
        
        
        if devType == '': devType = 'generic_termserver'
        if devType == '?': devType = ''
        

        
        sshConf = {
        'device_type': devType,
        'ip': list(hostCmdsList[0].keys())[0],
        'username': sshUser,
        'password': sshPass}
        
        print("Testing creds on first host in list: " + list(hostCmdsList[0].keys())[0])
        
        try:
            netmiko.ConnectHandler(**sshConf)
            print("Login success")
            return {'uid': sshUser, 'password': sshPass, 
                    'devType': devType}
            break
    
        except Exception as e:
            print(e)

        else:
            print('Access denied')
    
def runCommands(uid, password, devType):
    
    # Log File
    f = open("sshLog-{}.txt".format(timestamp), "a")
    
    for devs in hostCmdsList:
    
        for host, cmds in devs.items():
            
            sshConf = {
                'device_type': devType,
                'ip': host,
                'username': uid,
                'password': password}
            
            
            net_connect = netmiko.ConnectHandler(**sshConf)
            
            f.write("\n" + host)     
                
            sshOutput = net_connect.send_config_set(cmds)
            f.write(sshOutput)
            
            f.write("\n")
    
    f.close()
    


def main():
    
    createDictFromXls()
    
    credsDict = getCreds()
    
    runCommands(credsDict['uid'], credsDict['password'],
                 credsDict["devType"])

  
if __name__ == "__main__":
    main()